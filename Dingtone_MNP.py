import sys
import time
import requests
from PyQt6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QPushButton, QTextEdit, QTableWidget, QTableWidgetItem,
    QHBoxLayout, QFileDialog, QLabel, QHeaderView, QProgressBar, QMessageBox
)
from PyQt6.QtGui import QColor, QPalette
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains

OPERATOR_MAP = {
    "RJ": ("Reliance Jio", "https://imagerepo.ding.com/logo/RJ/IN.png"),
    "BL": ("BSNL", "https://imagerepo.ding.com/logo/BL/IN.png"),
    "VF": ("Vodafone", "https://imagerepo.ding.com/logo/VF/IN.png"),
    "AI": ("Airtel", "https://imagerepo.ding.com/logo/AI/IN.png")
}

def detect_operator_from_src(src: str) -> tuple:
    for key, (name, logo_url) in OPERATOR_MAP.items():
        if f"/{key}/" in src:
            return name, logo_url
    return "Not Found", ""

class OperatorChecker(QThread):
    result_signal = pyqtSignal(int, str, str, str)
    done_signal = pyqtSignal()
    progress_signal = pyqtSignal(int)
    current_number_signal = pyqtSignal(str)

    def __init__(self, numbers):
        super().__init__()
        self.numbers = numbers

    def normalize_number(self, num: str) -> str:
        digits = ''.join(filter(str.isdigit, num))
        if len(digits) == 10:
            return '+91' + digits
        elif len(digits) == 12 and digits.startswith('91'):
            return '+' + digits
        elif len(digits) == 13 and digits.startswith('91'):
            return '+' + digits[1:]
        return num

    def strip_country_code(self, num: str) -> str:
        return num[3:] if num.startswith('+91') else num

    def run(self):
        total = len(self.numbers)
        for index, number in enumerate(self.numbers):
            self.current_number_signal.emit(number)
            try:
                options = webdriver.ChromeOptions()
                options.add_argument("--disable-gpu")
                options.add_argument("--start-minimized")  # Minimized window
                driver = webdriver.Chrome(service=Service(), options=options)
                driver.set_window_position(-32000, 0)  # Move off-screen (Windows only)

                wait = WebDriverWait(driver, 30)
                driver.get("https://www.ding.com")

                try:
                    wait.until(EC.element_to_be_clickable((By.ID, "onetrust-accept-btn-handler"))).click()
                except:
                    pass

                input_box = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "input[data-testid='item-input']")))
                ActionChains(driver).move_to_element(input_box).click().perform()
                input_box.clear()
                input_box.send_keys(number)

                start_btn = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button[data-testid='button-country-widget']")))
                driver.execute_script("arguments[0].click();", start_btn)

                logo = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "img[data-testid='summary-product-operator-logo']")))
                src = logo.get_attribute("src")

                operator, logo_url = detect_operator_from_src(src)
                display_number = self.strip_country_code(number)
                self.result_signal.emit(index + 1, display_number, operator, logo_url)
            except Exception:
                display_number = self.strip_country_code(number)
                self.result_signal.emit(index + 1, display_number, "Error", "")
            finally:
                self.progress_signal.emit(int(((index + 1) / total) * 100))
                driver.quit()

        self.done_signal.emit()

class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("📶 Ding Operator Checker")
        self.resize(1000, 750)
        palette = self.palette()
        palette.setColor(QPalette.ColorRole.Window, QColor("#121212"))
        palette.setColor(QPalette.ColorRole.WindowText, QColor("white"))
        self.setPalette(palette)

        layout = QVBoxLayout()

        self.input_box = QTextEdit()
        self.input_box.setPlaceholderText("Enter one number per line...")
        self.input_box.setFixedHeight(100)
        self.input_box.setStyleSheet("background-color: #1e1e1e; color: white; font-size: 14px; border: 1px solid #444;")

        self.current_label = QLabel("📱 Currently Searching: -")
        self.current_label.setStyleSheet("color: cyan; font-weight: bold; padding: 4px;")

        btn_layout = QHBoxLayout()
        self.start_btn = QPushButton("🔍 Start")
        self.import_btn = QPushButton("📂 Import")
        self.export_btn = QPushButton("📅 Export")
        self.copy_btn = QPushButton("📋 Copy")
        self.paste_btn = QPushButton("📥 Paste")
        self.clear_btn = QPushButton("🩹 Clear")
        self.copy_results_btn = QPushButton("📄 Copy Results")

        for btn in [self.start_btn, self.import_btn, self.export_btn, self.copy_btn, self.paste_btn, self.clear_btn, self.copy_results_btn]:
            btn.setStyleSheet("""
                QPushButton {
                    background-color: #333;
                    color: white;
                    font-weight: bold;
                    padding: 8px 12px;
                    border-radius: 6px;
                }
                QPushButton:hover {
                    background-color: #555;
                }
            """)
            btn_layout.addWidget(btn)

        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        self.progress_bar.setStyleSheet("QProgressBar { color: white; text-align: center; background-color: #2a2a2a; border: 1px solid #555; border-radius: 5px; } QProgressBar::chunk { background-color: #00bc8c; }")

        self.start_btn.clicked.connect(self.start_lookup)
        self.import_btn.clicked.connect(self.import_excel)
        self.export_btn.clicked.connect(self.export_excel)
        self.copy_btn.clicked.connect(self.copy_text)
        self.paste_btn.clicked.connect(self.paste_text)
        self.clear_btn.clicked.connect(self.clear_input)
        self.copy_results_btn.clicked.connect(self.copy_results)

        self.table = QTableWidget()
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["Sr. No.", "Number", "Operator"])
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.horizontalHeader().setStyleSheet("QHeaderView::section { background-color: #000000; color: #00ffff; font-weight: bold; padding: 6px; }")
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setStyleSheet("QTableWidget { background-color: #1c1c1c; color: white; gridline-color: #444; } QTableWidget::item:selected { background-color: #444; }")

        layout.addWidget(self.input_box)
        layout.addWidget(self.current_label)
        layout.addLayout(btn_layout)
        layout.addWidget(self.progress_bar)
        layout.addWidget(self.table)
        self.setLayout(layout)

    def start_lookup(self):
        raw = self.input_box.toPlainText()
        numbers = list(filter(None, {self.normalize_number(line.strip()) for line in raw.splitlines()}))

        self.table.setRowCount(0)
        self.progress_bar.setValue(0)

        self.worker = OperatorChecker(numbers)
        self.worker.result_signal.connect(self.add_result)
        self.worker.progress_signal.connect(self.progress_bar.setValue)
        self.worker.done_signal.connect(self.search_finished)
        self.worker.current_number_signal.connect(self.update_current_label)
        self.worker.start()
        self.start_btn.setEnabled(False)

    def update_current_label(self, number):
        self.current_label.setText(f"📱 Currently Searching: {number}")

    def normalize_number(self, num: str) -> str:
        digits = ''.join(filter(str.isdigit, num))
        if len(digits) == 10:
            return '+91' + digits
        elif len(digits) == 12 and digits.startswith('91'):
            return '+' + digits
        elif len(digits) == 13 and digits.startswith('91'):
            return '+' + digits[1:]
        return num

    def add_result(self, sr_no, number, operator, logo_url):
        row = self.table.rowCount()
        self.table.insertRow(row)
        self.table.setItem(row, 0, QTableWidgetItem(str(sr_no)))
        self.table.setItem(row, 1, QTableWidgetItem(number))
        self.table.setItem(row, 2, QTableWidgetItem(operator))
        for col in range(3):
            item = self.table.item(row, col)
            if item:
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

    def import_excel(self):
        try:
            import openpyxl
        except ImportError:
            QMessageBox.critical(self, "Missing Dependency", "Install openpyxl to import Excel files.")
            return

        path, _ = QFileDialog.getOpenFileName(self, "Import Excel", "", "Excel Files (*.xlsx *.xls)")
        if path:
            from openpyxl import load_workbook
            wb = load_workbook(path)
            ws = wb.active
            numbers = [str(cell.value).strip() for cell in ws['A'] if cell.value]
            self.input_box.setPlainText("\n".join(numbers))

    def export_excel(self):
        try:
            import openpyxl
        except ImportError:
            QMessageBox.critical(self, "Missing Dependency", "Install openpyxl to export Excel files.")
            return

        path, _ = QFileDialog.getSaveFileName(self, "Export Excel", "", "Excel Files (*.xlsx)")
        if path:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.append(["Number", "Operator"])
            for row in range(self.table.rowCount()):
                number = self.table.item(row, 1).text()
                operator = self.table.item(row, 2).text()
                ws.append([number, operator])
            wb.save(path)

    def copy_text(self):
        clipboard = QApplication.clipboard()
        clipboard.setText(self.input_box.toPlainText())

    def paste_text(self):
        self.input_box.paste()

    def clear_input(self):
        self.input_box.clear()

    def copy_results(self):
        text = "Sr. No.\tNumber\tOperator\n"
        for row in range(self.table.rowCount()):
            sr = self.table.item(row, 0).text()
            number = self.table.item(row, 1).text()
            operator = self.table.item(row, 2).text()
            text += f"{sr}\t{number}\t{operator}\n"
        QApplication.clipboard().setText(text)
        QMessageBox.information(self, "Copied", "✅ Results copied to clipboard!")

    def search_finished(self):
        self.start_btn.setEnabled(True)
        self.current_label.setText("📱 Currently Searching: -")
        QMessageBox.information(self, "Completed", "✅ Search complete!")

if __name__ == '__main__':
    app = QApplication(sys.argv)
    win = MainWindow()
    win.show()
    sys.exit(app.exec())
